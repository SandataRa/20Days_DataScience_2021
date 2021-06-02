#Day 17

# Constructing Data Visualizations

import openpyxl
from openpyxl.chart import LineChart, Reference

wb = openpyxl.load_workbook('AirLineBumps_16_17.xlsx')
sheet = wb['timeanalysis']

values = Reference(sheet, min_col = 2, min_row = 2, max_col = 2,
                  max_row = sheet.max_row)

chart = LineChart()

chart.add_data(values)
chart.title = "Airlines Bumps"

chart.x_axis.title = "Year"
chart.y_axis.title = "Number of bumps"

sheet.add_chart(chart, "E1")

wb.save('AirLineBumps_linechart.xlsx')

# Customizing Chart Appearance

chart = LineChart()

chart.title = "Airlines Bumps"
chart.x_axis.title = "Year"
chart.y_axis.title = "Number of bumps"



values = Reference(sheet, min_col = 2, min_row = 1, max_col = 2,
                  max_row = sheet.max_row)

chart.add_data(values, titles_from_data = True)
dates = Reference(sheet, min_col = 1, min_row = 2,
                  max_col = 1, max_row = sheet.max_row)

chart.set_categories(dates)

# chart.x_axis.tickLblSkip = 24
# chart.width = 25
# chart.height = 12.5

sheet.add_chart(chart, "E16")

wb.save('AirLineBumps_linechart.xlsx')
from openpyxl.chart import AreaChart

area_chart = AreaChart()

area_chart.title = "AirLine Bumps"
area_chart.x_axis.title = "Year"
area_chart.y_axis.title = "Number of Bumps"

area_chart.add_data(values, titles_from_data = True)

area_chart.set_categories(dates)

area_chart.series

area_chart.series[0].graphicalProperties.solidFill = "ff9900"

sheet.add_chart(area_chart, "E31")

wb.save('AirLineBumps_linechart.xlsx')

wb = openpyxl.load_workbook('AirLineBumps_16_17.xlsx')
sheet = wb['timeanalysis']

multi_line_chart = LineChart()

multi_line_chart.title = "Delta Air Lines vs Virgin America"
multi_line_chart.x_axis.title = "Year"
multi_line_chart.y_axis.title = "Number of bumps"

values = Reference(sheet, min_col = 2, min_row = 1, max_col = 3,
                  max_row = sheet.max_row)

multi_line_chart.add_data(values, titles_from_data = True)

dates = Reference(sheet, min_col = 1, min_row = 2,
                  max_col = 1, max_row = sheet.max_row)


multi_line_chart.set_categories(dates)

len(multi_line_chart.series)

delta_series  = multi_line_chart.series[0]
virgin_series = multi_line_chart.series[1]

delta_series.graphicalProperties.line.solidFill = '000000'
delta_series.graphicalProperties.line.width = 2 

virgin_series.graphicalProperties.line.solidFill = '0000FF'
virgin_series.graphicalProperties.line.dashStyle = 'sysDot' 

sheet.add_chart(multi_line_chart, "E1")

wb.save('AirLineBumps_multi_line_chart.xlsx')


from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

work_book = openpyxl.load_workbook('AirLineBumps_16_17.xlsx')
sheet = work_book['dataset']

chart = BarChart()
chart.type = "col"
chart.title = "Bar Chart"
chart.x_axis.title = "Bumps per Airlines Company"
chart.y_axis.title = "Bumps / 10K passengers"
chart.width = 10
chart.height = 7.5
data = Reference(worksheet = sheet, min_row = 1,
                max_row = sheet.max_row, min_col = 2,
                max_col = 3)
cats = Reference(worksheet = sheet, min_row = 2,
                max_row = sheet.max_row, min_col = 1,
                max_col = 1)


chart.add_data(data, titles_from_data = True)
chart.set_categories(cats)
sheet.add_chart(chart, "H1")


work_book.save('AirLineBumps_bar_chart.xlsx')


from copy import deepcopy


hor_chart = deepcopy(chart)

hor_chart.type = "bar"
hor_chart.title = "Horizontal Bar Chart"
sheet.add_chart(hor_chart, "H16")
work_book.save('AirLineBumps_bar_chart.xlsx')
stacked_chart = deepcopy(chart)
stacked_chart.type =  "col"
stacked_chart.grouping = "stacked"
stacked_chart.overlap = 100
stacked_chart.title = "Stacked Chart"
sheet.add_chart(stacked_chart, "N1")
work_book.save('AirLineBumps_bar_chart.xlsx')
from openpyxl.chart import BarChart3D
chart_3d = BarChart3D()
chart_3d.type = "col"
chart_3d.title = "3D Bar Chart"
chart_3d.y_axis.title = "Number of bumps"
chart_3d.x_axis.title="Airline"
chart_3d.add_data(data, titles_from_data = True)
chart_3d.set_categories(cats)
sheet.add_chart(chart_3d, "N16")
work_book.save('AirLineBumps_bar_chart.xlsx')
