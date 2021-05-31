#Day 15

import openpyxl
work_book = openpyxl.load_workbook('AirLineBumps_16_17.xlsx')
work_book.sheetnames
sheet_obj = work_book['dataset']
sheet_obj['A2'].value +  ' has bumped ' + str(sheet_obj['B2'].value) + ' passengers in September 2017'

max_row = sheet_obj.max_row

for i in range(2, max_row-1):
    airline = sheet_obj.cell(row = i, column = 1).value
    bumps_17 = str(sheet_obj.cell(row = i, column = 2).value)
    print(airline + ' bumped ' + bumps_17 + ' passengers in September 2017.')

import json

bumps = {}

for row in sheet_obj.iter_rows(min_row = 2,
                              max_row = 4,
                              min_col = 1,
                              max_col = 3,
                              values_only = True):
    airline = row[0]
    b_details = {
        "2017": row[1],
        "2016": row[2]
    }
    bumps[airline] = b_details

from openpyxl.styles import numbers
for row in sheet['B2:C13']:
    for cell in row: 
        cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

yellow_background = PatternFill(bgColor = colors.COLOR_INDEX[5])

diff_style = DifferentialStyle(fill = yellow_background)

rule = Rule(type="expression", dxf = diff_style)

rule.formula = ["$B1<1000"]

sheet.conditional_formatting.add(sheet.calculate_dimension(), rule)

work_book.save('AirLineBumps_16_17_formatted.xlsx')

# Analyzing the potential causes of attrition among the Research & Development department 
f2 = (dataset_subset.Department == 'Research & Development')

rd_dataset = dataset_subset[f2]

# Building the dashboard of potential causes of attritions
f, axes = plt.subplots(2,3, figsize = (15,15))

def plotStackedHistoSubplot (ax, data, xaxis, title):
    plt.rcParams['figure.figsize'] = 8,4
    sns.set_style("darkgrid")
    
    filters = []
    labels = list()
    for g in data.Attrition.cat.categories :
        labels.append(g)
        elem = data[data.Attrition == g][xaxis]
        filters.append(elem)
    plt.xlabel(xaxis, fontsize = 10, color = "Black")
    plt.title(title, fontsize = 15, color="Black", \
         fontname = "Arial")
    ax.set_title(title)
    ax.set(xlabel=xaxis)    
    ax.hist(np.array(filters, dtype='object'), bins = 30, stacked = True, label = labels)
    ax.plot()
    ax.legend(loc='upper right')
    
k1 = plotStackedHistoSubplot (axes[0,0], rd_dataset, 'PerformanceRating', "Performance Ratings by Attrition")
k2 = plotStackedHistoSubplot (axes[0,1], rd_dataset, 'WorkLifeBalance', "Work Life Balance by Attrition by Attrition")
k3 = plotStackedHistoSubplot (axes[0,2], rd_dataset, 'YearsSinceLastPromotion', "Year Since Last Promotion by Attrition")

k4 = plotStackedHistoSubplot (axes[1,0], rd_dataset, 'TrainingTimesLastYear', "Training Times Last Year by Attrition")
k5 = plotStackedHistoSubplot (axes[1,1], rd_dataset, 'YearsWithCurrManager', "Years with Current Manager by Attrition")
k6 = plotStackedHistoSubplot (axes[1,2], rd_dataset, 'StockOptionLevel', "Stock option level by Attrition")





